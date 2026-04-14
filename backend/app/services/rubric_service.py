"""Rubric service - AI analysis and chat for rubric evaluation."""

from __future__ import annotations

import io
import json
import logging
import re
from time import perf_counter
from typing import Any

import httpx
from fastapi import HTTPException

from app.core.config import settings
from app.schemas.rubric import ChatMessage, RubricAnalysis, RubricItem


logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────
# HTTP Client singleton
# ──────────────────────────────────────────────────────────────

_http_client: httpx.AsyncClient | None = None


async def get_http_client() -> httpx.AsyncClient:
    """Get or create HTTP client singleton."""
    global _http_client
    if _http_client is None or _http_client.is_closed:
        _http_client = httpx.AsyncClient(
            timeout=httpx.Timeout(connect=10.0, read=120.0, write=30.0, pool=5.0),
            limits=httpx.Limits(max_connections=100, max_keepalive_connections=20),
        )
    return _http_client


async def close_http_client() -> None:
    """Close HTTP client on shutdown."""
    global _http_client
    if _http_client is not None and not _http_client.is_closed:
        await _http_client.aclose()
        _http_client = None


# ──────────────────────────────────────────────────────────────
# Shared helpers
# ──────────────────────────────────────────────────────────────


def _strip_think_tags(text: str) -> str:
    """Keep only content after </think>; return text as-is if tag absent."""
    marker = "</think>"
    idx = text.find(marker)
    return text[idx + len(marker) :].strip() if idx != -1 else text.strip()


def _apply_thinking_control(payload: dict[str, Any]) -> dict[str, Any]:
    payload["chat_template_kwargs"] = {
        **dict(payload.get("chat_template_kwargs") or {}),
        "enable_thinking": settings.TEMPLATE_RECOMMENDATION_VLLM_ENABLE_THINKING,
    }
    return payload


def _vllm_headers() -> dict[str, str]:
    return {
        "Authorization": f"Bearer {settings.TEMPLATE_RECOMMENDATION_VLLM_API_KEY}",
        "Content-Type": "application/json",
    }


def _to_float(value: Any, default: float = 0.0) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return default
        try:
            return float(text)
        except ValueError:
            match = re.search(r"-?\d+(?:\.\d+)?", text)
            if match:
                try:
                    return float(match.group(0))
                except ValueError:
                    return default
    return default


def _normalize_rubric_items(raw_items: Any) -> list[RubricItem]:
    """Best-effort normalization for AI-returned item payloads."""
    if not isinstance(raw_items, list):
        return []

    normalized: list[RubricItem] = []
    for i, raw in enumerate(raw_items):
        if not isinstance(raw, dict):
            continue

        item_id = str(raw.get("id") or f"item-{i + 1}")
        title = str(raw.get("title") or raw.get("name") or "").strip() or "未命名項目"
        description = str(raw.get("description") or raw.get("desc") or "")
        max_score = _to_float(raw.get("max_score", raw.get("score", 0.0)))

        detectable = str(raw.get("detectable") or "manual").strip().lower()
        if detectable not in {"auto", "partial", "manual"}:
            detectable = "manual"

        detection_method = raw.get("detection_method") or raw.get("detection")
        fallback = raw.get("fallback") or raw.get("suggestion")

        normalized.append(
            RubricItem(
                id=item_id,
                title=title,
                description=description,
                max_score=max_score,
                detectable=detectable,
                detection_method=str(detection_method)
                if detection_method is not None
                else None,
                fallback=str(fallback) if fallback is not None else None,
            )
        )

    return normalized


def normalize_items_for_export(raw_items: Any) -> list[RubricItem]:
    """Public helper for robust export parsing."""
    return _normalize_rubric_items(raw_items)


def _extract_context_item_count(rubric_context: str) -> int:
    try:
        parsed = json.loads(rubric_context or "{}")
    except json.JSONDecodeError:
        return 0
    items = parsed.get("items")
    return len(items) if isinstance(items, list) else 0


async def _call_vllm(
    payload: dict[str, Any], timeout: float = 60.0
) -> tuple[str, dict]:
    """Call vLLM chat/completions and return (content, usage_metrics)."""
    url = f"{settings.TEMPLATE_RECOMMENDATION_VLLM_BASE_URL}/chat/completions"
    started = perf_counter()

    logger.debug(f"Calling vLLM API: {url}")

    try:
        client = await get_http_client()
        resp = await client.post(
            url,
            json=payload,
            headers=_vllm_headers(),
            timeout=timeout,
        )
        resp.raise_for_status()
        data = resp.json()

        elapsed = max(perf_counter() - started, 0.0)
        usage = data.get("usage") or {}
        prompt_tokens = int(usage.get("prompt_tokens") or 0)
        completion_tokens = int(usage.get("completion_tokens") or 0)
        total_tokens = int(
            usage.get("total_tokens") or (prompt_tokens + completion_tokens)
        )
        tps = (completion_tokens / elapsed) if elapsed > 0 else 0.0

        logger.info(
            f"vLLM call successful: {total_tokens} tokens in {elapsed:.2f}s ({tps:.1f} t/s)"
        )

        content = data["choices"][0]["message"]["content"] or ""
        content = _strip_think_tags(content)
        metrics = {
            "prompt_tokens": prompt_tokens,
            "completion_tokens": completion_tokens,
            "total_tokens": total_tokens,
            "elapsed_seconds": round(elapsed, 3),
            "tokens_per_second": round(tps, 2),
        }
        return content, metrics
    except httpx.TimeoutException as exc:
        logger.error(f"vLLM API timeout after {timeout}s")
        raise HTTPException(
            status_code=504, detail="AI 服務回應超時，請稍後再試。"
        ) from exc
    except httpx.HTTPStatusError as exc:
        status = exc.response.status_code
        logger.error(f"vLLM API returned status {status}")
        raise HTTPException(
            status_code=502, detail=f"AI 服務異常（狀態碼 {status}）"
        ) from exc
    except Exception as exc:
        logger.error(f"vLLM API call failed: {exc}", exc_info=True)
        raise HTTPException(status_code=502, detail=f"AI 呼叫失敗：{exc}") from exc


# ──────────────────────────────────────────────────────────────
# 1. Analyze rubric document
# ──────────────────────────────────────────────────────────────

_ANALYZE_SYSTEM_PROMPT = """
# 角色
你是一位專業的教學評分助理，服務對象是校園雲端平台的授課老師。

# 平台背景（內部知識）
本系統為校園雲端平台，學生作業在以下環境中執行：
- Proxmox 虛擬機器（VM）或 LXC 容器
- 本地部署，不使用公有雲（AWS/GCP/Azure）

## 系統可以自動偵測的資訊（透過 Proxmox Agent / API）
- 特定 TCP Port 是否正在監聽（e.g., Port 80/443/3306/5432）
- Linux 服務狀態（`systemctl status nginx` 等）
- 進程是否存在（`ps aux`）
- 磁碟/CPU/記憶體使用率
- 特定檔案是否存在（透過 Agent Exec）
- HTTP endpoint 回傳狀態碼

## 系統無法自動偵測的資訊（需人工判斷）
- 程式碼品質、架構設計、MVC 分層等
- 資料庫內容正確性（需帳密才能查詢）
- 圖形化介面截圖（需 VNC/螢幕截圖）
- Docker Compose 或設定檔內容（需讀原始碼，須學生授權）
- 報告、文件、簡報品質
- 功能邏輯正確性（需 E2E 操作測試）

# 任務
根據以下評分表原始文字，完成兩件事：
1. 萃取所有評分項目，轉為 JSON 列表。
2. 針對每一個評分項目，依據上述平台能力，判斷其「可自動偵測性」：
   - "auto"：可完全透過系統自動偵測，無需人工介入
   - "partial"：需要部分人工輔助或額外授權才能偵測
   - "manual"：完全需要人工評閱，系統無法偵測

# 輸出格式
只輸出合法的 JSON，不要有任何說明文字或 markdown。結構如下：
{
  "items": [
    {
      "id": "item-1",
      "title": "評分項目名稱",
      "description": "評分說明（從原文萃取或精簡改寫）",
      "max_score": 數字,
      "detectable": "auto | partial | manual",
      "detection_method": "若 auto/partial，具體說明偵測方式（e.g., TCP Port 80 探測）；否則 null",
      "fallback": "若 manual/partial，說明替代方案（e.g., 請學生提交截圖、要求提交 GitHub 連結）；否則 null"
    }
  ],
  "summary": "整體評分表說明，約 2-3 句，使用繁體中文。包含總配分、可自動偵測比例、哪些項目需特別注意或無法自動評分。"
}
""".strip()


async def analyze_rubric(raw_text: str) -> tuple[RubricAnalysis, dict]:
    """Send raw document text to AI, return structured RubricAnalysis."""
    if not settings.TEMPLATE_RECOMMENDATION_VLLM_MODEL_NAME:
        raise HTTPException(
            status_code=503, detail="TEMPLATE_RECOMMENDATION_VLLM_MODEL_NAME 未設定。"
        )

    logger.info(f"Starting rubric analysis, text length: {len(raw_text)} characters")

    user_content = f"# 評分表原文\n\n{raw_text}"

    payload = _apply_thinking_control(
        {
            "model": settings.TEMPLATE_RECOMMENDATION_VLLM_MODEL_NAME,
            "messages": [
                {"role": "system", "content": _ANALYZE_SYSTEM_PROMPT},
                {"role": "user", "content": user_content},
            ],
            "max_tokens": settings.TEMPLATE_RECOMMENDATION_VLLM_MAX_TOKENS,
            "temperature": 0.2,
            "top_p": settings.TEMPLATE_RECOMMENDATION_VLLM_TOP_P,
            "response_format": {"type": "json_object"},
        }
    )

    content, metrics = await _call_vllm(
        payload, timeout=float(settings.TEMPLATE_RECOMMENDATION_VLLM_TIMEOUT)
    )

    try:
        data = json.loads(content)
    except json.JSONDecodeError as exc:
        logger.error(f"Failed to parse AI response as JSON: {exc}")
        raise HTTPException(
            status_code=502, detail=f"AI 回傳 JSON 解析失敗：{exc}"
        ) from exc

    # 使用統一的正規化函數處理 AI 回傳的項目
    items_raw = data.get("items") or []
    items = _normalize_rubric_items(items_raw)

    total_score = sum(item.max_score for item in items)
    auto_count = sum(1 for item in items if item.detectable == "auto")
    partial_count = sum(1 for item in items if item.detectable == "partial")
    manual_count = sum(1 for item in items if item.detectable == "manual")

    logger.info(
        f"Analysis complete: {len(items)} items, {total_score} total points (auto: {auto_count}, partial: {partial_count}, manual: {manual_count})"
    )

    analysis = RubricAnalysis(
        items=items,
        total_score=total_score,
        auto_count=auto_count,
        partial_count=partial_count,
        manual_count=manual_count,
        summary=str(data.get("summary") or ""),
        raw_text=raw_text,
    )
    return analysis, metrics


# ──────────────────────────────────────────────────────────────
# 2. Chat to refine rubric
# ──────────────────────────────────────────────────────────────

_CHAT_SYSTEM_TEMPLATE = """
# 角色
你是一位專業的教學評分助理，服務對象是校園雲端平台的授課老師。
老師已上傳了一份評分表，你已完成初步分析。

# 平台背景（內部知識）
學生作業在 Proxmox VM / LXC 環境中執行（本地校園雲端，非公有雲）。
系統可偵測：Port 監聽、服務狀態、CPU/記憶體/磁碟、HTTP 狀態碼、檔案是否存在。
系統無法偵測：程式碼品質、DB 內容、設定檔內容、截圖、報告品質。

# 可用資訊來源
- 評分表結構（見下方 JSON）
- 未來將支援讀取學生 README、程式碼片段以輔助判斷（功能開發中）

# 目前評分表（JSON 格式）
{rubric_context}

# 目前評分項目總數
{rubric_item_count}

# 情境說明與任務
{situation_instruction}

# 輸出規則（不論任何情境一律遵守）

## 對話風格
- 使用**親切、專業但有溫度**的繁體中文白話文。
- 想像你是老師的教學助理，用「我」稱呼自己，用「你」或「您」稱呼老師（視語境自然選擇）。
- 避免過度正式的用語（例如「依據分析」「經判斷」），改用自然對話（例如「我看了一下」「這個的話」）。
- 不得提到技術欄位名稱（id、detectable、detection_method、fallback 等）。
- 不得在 reply 中重新列出所有項目或逐項說明修改細節。

## 精簡回覆原則（極為重要）
- 當執行修改時，只需**簡短總結**變更結果，不要逐一列出每個欄位的前後差異。
- 修改應在背景靜默完成，回覆只需告知老師「做了什麼」和「結果如何」。
- 好的回覆範例：「好的，已幫你調整了 3 個項目的偵測方式，配分也一併更新了。」
- 不好的回覆範例：「已修改如下：第 1 項：標題從 X 改為 Y，配分從 5 改為 10，偵測方式從...改為...；第 2 項：...」

## 語氣示範
❌ 不好：「經分析，該項目不符合自動偵測條件，建議歸類為手動評閱。」
✅ 良好：「這個項目需要檢查程式碼品質，目前系統還做不到自動判斷，建議由你親自評閱會比較準確。」

❌ 不好：「已將項目 3 的可偵測性修改為自動。」
✅ 良好：「好的！我已經把第 3 項改成可以自動偵測了。」

❌ 不好：「根據您的需求，更新如下...」
✅ 良好：「了解！我幫你調整好了，主要改了這幾個地方...」

## 不確定性表達
- 當問題涉及**邊界情境**（需要額外授權、學生配合、或技術上可行但不建議）時，請明確說明限制條件，並提供多種方案供老師選擇，不要一口咬定「可以」或「不行」。
- 當你真的不確定時，誠實說「這個我需要再確認一下」或「這個情況比較複雜」，不要硬給答案。

## 不確定性示範
❌ 不好：「系統可以自動檢查資料庫內容。」
✅ 良好：「這個需要學生在設定檔提供資料庫帳密，系統才能自動連線檢查。但這樣可能有資安疑慮，而且設定過程學生容易出錯。我建議改為『檢查資料庫服務是否啟動』（這個可以自動偵測），內容正確性的部分由你親自評閱。你覺得如何？」

# 輸出格式
只輸出合法 JSON，不要任何 markdown 包裹或自然語言。結構：
{{
  "reply": "你的白話回覆",
  "updated_items": null
}}

IMPORTANT RULES - 意圖識別（極為重要）

請仔細判斷老師的意圖，區分「詢問/討論」和「執行指令」：

## 【只聊天不改表單】（updated_items 必須為 null）
以下情況只提供建議，不修改評分表：
1. **詢問性問句**：句子有問號，且在徵求意見
   - 例如：「這個可以自動偵測嗎？」「你覺得呢？」「建議怎麼改？」
2. **試探性討論**：老師在討論可能性，未下達指令
   - 例如：「如果改成...會不會比較好？」「這樣可以嗎？」「要不要把...」
3. **純粹解釋**：詢問原因或要求說明
   - 例如：「為什麼這項是人工評閱？」「有其他方案嗎？」「怎麼判斷的？」
4. **要求建議但未明確執行**：只要建議，沒說「照做」
   - 例如：「給我一些建議」「分析一下」「看看有什麼問題」

## 【需要修改表單】（updated_items 必須為完整列表）
以下情況才執行修改：
1. **明確指令動詞**：有清楚的行動指示
   - 例如：「請修改」「幫我改」「新增」「刪除」「調整為」「改成」
2. **確認執行**：老師對前面的建議表示同意並要執行
   - 例如：「好」「OK」「就這樣做」「照你說的改」「可以，請執行」
3. **直接陳述變更**：直接說要改什麼，沒有問號
   - 例如：「第 3 項改成自動偵測」「把配分改成 10 分」「刪掉第 5 項」

## 黃金原則
**當你無法確定是「詢問」還是「指令」時，優先視為「詢問」，僅提供建議。**
如果老師滿意你的建議，會明確說「好，請修改」或「就這樣做」。

## 輸出結構約束
- 每個 item 需有：id, title, description, max_score, detectable, detection_method, fallback。
- 若修改表單（updated_items 不為 null），必須包含「完整列表」，未被要求變更的項目必須原樣保留，不得省略。

## ⚠️ 極為重要：完整列表規則
當你需要修改評分表時（updated_items 不為 null）：
1. **必須返回完整的 {rubric_item_count} 個項目**，一個都不能少
2. 只修改老師要求變更的項目，其他項目完全照抄原本的資料
3. 即使老師只說「把第 3 項改成自動偵測」，你也要返回所有 {rubric_item_count} 個項目
4. 如果老師要求刪除項目，返回的項目數會減少，這是唯一允許項目數變少的情況
5. 如果老師要求新增項目，返回的項目數會增加

**錯誤範例**（絕對不可以這樣做）：
- 老師說「把第 3 項配分改成 10」→ 你只返回第 3 項 ❌
- 老師說「新增一個檢查 Port 80 的項目」→ 你只返回新增的項目 ❌

**正確範例**：
- 老師說「把第 3 項配分改成 10」→ 你返回全部項目，只有第 3 項的 max_score 改成 10 ✅
- 老師說「新增一個檢查 Port 80 的項目」→ 你返回原本所有項目 + 新增的項目 ✅
""".strip()


_SITUATION_NORMAL = """
老師正在對話中修改或詢問評分表相關問題。

## 你的任務
- 老師可能會**詢問**特定項目的偵測方式、配分建議、可行性分析。
- 老師也可能會**下達指令**，要求修改、新增、刪除評分項目。
- 請依據上方的「意圖識別規則」判斷老師是在詢問還是下指令。

## 處理原則
- 若老師詢問的項目涉及無法自動偵測的內容，請主動說明限制並給出替代方案。
- 提供建議時，語氣要自然親切，可以用「我建議...」「你覺得...如何？」等句式。
- 只有在老師明確下達指令時，才修改評分表（updated_items 不為 null）。
""".strip()

_SITUATION_REFINE = """
老師剛剛親手調整了評分表，現在請你進行「全表審核潤飾」。

## 你的任務（按優先順序執行）

### 1. 審核一致性（最重要）
檢查每個項目的「可偵測性判斷」與「檢查方式」是否邏輯一致：
- 例如：標記為可自動偵測，但檢查方式是「人工檢視程式碼」→ 不一致，需更正。
- 例如：標記為人工評閱，但檢查方式是「偵測 Port 80」→ 不一致，需更正。

### 2. 補齊空白欄位
針對未填寫的「偵測方式」(detection_method) 和「替代建議」(fallback)，依平台能力推斷並補充：
- **重要**：若老師已填寫，絕不修改，除非明顯與可偵測性矛盾。
- 只補充「空白」或「null」的欄位。

### 3. 語氣統一與明確化（保守原則）
只潤飾以下情況的項目：
- 說明過於簡略（例如只有 2-3 個字，看不懂在說什麼）。
- 語氣明顯不統一（例如有些用「學生須...」，有些用「請檢查...」）。
- 有明顯的錯字或語病。

**絕對不要**：
- 不要自作主張改寫老師的專業術語（例如老師寫「LAMP 架構」就保留，不要改成「Linux + Apache + MySQL + PHP」）。
- 不要改動老師仔細填寫過的完整說明文字。
- 不要為了「統一風格」而大幅改寫原文。

## 保守原則（極為重要）
**當你不確定是否該修改時，保持原樣。**
只修正明顯的錯誤（矛盾、空白、語病），不做「美化」或「風格統一」的過度編輯。

## 回覆要求（精簡總結，不列細節）
- 只需**一句話總結**你做了哪些調整（例如「審核完畢，共調整了 X 個項目的偵測判斷，並補齊了 Y 處空白欄位。」）。
- 如果有特別重要的矛盾或需要老師留意的地方，可簡短提及，但不要逐項列出所有修改。
- 如果沒有需要調整的地方，簡短說明「檢查完畢，評分表目前狀態良好。」
- 修改在背景完成即可，老師可以直接在表單上看到變更結果。
""".strip()


async def chat_with_rubric(
    messages: list[ChatMessage],
    rubric_context: str,
    is_refine: bool = False,
) -> tuple[str, list | None, dict]:
    """
    Multi-turn chat with rubric context injected into system prompt.
    Returns (reply_text, updated_items_or_None, metrics).
    - is_refine: True 表示老師手動修改完表單後觸發的「全表潤飾」模式。
    - updated_items: complete list of RubricItem dicts when AI modified the rubric;
      None when AI only answered a question without changes.
    """
    if not settings.TEMPLATE_RECOMMENDATION_VLLM_MODEL_NAME:
        raise HTTPException(
            status_code=503, detail="TEMPLATE_RECOMMENDATION_VLLM_MODEL_NAME 未設定。"
        )

    context_item_count = _extract_context_item_count(rubric_context)
    situation = _SITUATION_REFINE if is_refine else _SITUATION_NORMAL
    system_prompt = (
        _CHAT_SYSTEM_TEMPLATE.replace(
            "{rubric_context}", rubric_context or "（尚未上傳評分表）"
        )
        .replace("{rubric_item_count}", str(context_item_count))
        .replace("{situation_instruction}", situation)
    )

    formatted = [{"role": "system", "content": system_prompt}]
    for msg in messages:
        formatted.append({"role": msg.role, "content": msg.content})

    payload = _apply_thinking_control(
        {
            "model": settings.TEMPLATE_RECOMMENDATION_VLLM_MODEL_NAME,
            "messages": formatted,
            "max_tokens": settings.TEMPLATE_RECOMMENDATION_VLLM_CHAT_MAX_TOKENS,
            "temperature": settings.TEMPLATE_RECOMMENDATION_VLLM_CHAT_TEMPERATURE,
            "top_p": settings.TEMPLATE_RECOMMENDATION_VLLM_TOP_P,
            "top_k": settings.TEMPLATE_RECOMMENDATION_VLLM_TOP_K,
            "repetition_penalty": settings.TEMPLATE_RECOMMENDATION_VLLM_REPETITION_PENALTY,
            "response_format": {"type": "json_object"},
        }
    )

    content, metrics = await _call_vllm(
        payload, timeout=float(settings.TEMPLATE_RECOMMENDATION_VLLM_TIMEOUT)
    )

    # 解析結構化 JSON 回覆
    reply_text = content  # fallback
    updated_items: list | None = None
    try:
        parsed = json.loads(content)
        reply_text = str(parsed.get("reply") or content)
        raw_updated = parsed.get("updated_items")
        normalized_updated = _normalize_rubric_items(raw_updated)
        if normalized_updated:
            # ⚠️ 安全檢查：防止 AI 返回不完整的列表
            # 檢查 AI 返回的項目數量是否異常減少
            if context_item_count > 0:
                updated_count = len(normalized_updated)
                # 如果項目數減少超過 1 個，且不是精煉模式（精煉模式不應該刪除項目）
                if updated_count < context_item_count - 1:
                    logger.warning(
                        f"⚠️ AI 返回的項目數異常：期望至少 {context_item_count - 1} 個，"
                        f"實際返回 {updated_count} 個。可能導致資料遺失。"
                    )
                    # 在回覆中加入警告訊息
                    reply_text = (
                        f"⚠️ 系統偵測到異常：我只返回了 {updated_count} 個項目，"
                        f"但原本有 {context_item_count} 個。這可能是我理解錯誤了。\n\n"
                        f"為了安全起見，請確認這是否是你想要的結果。如果不是，請重新說明你的需求。\n\n"
                        f"原始回覆：{reply_text}"
                    )
            updated_items = [item.model_dump() for item in normalized_updated]
    except (json.JSONDecodeError, TypeError):
        pass  # AI 未輸出合法 JSON，直接用原始內容作為回覆

    return reply_text, updated_items, metrics


# ──────────────────────────────────────────────────────────────
# 3. Export to Excel
# ──────────────────────────────────────────────────────────────

_DETECTABLE_LABELS = {
    "auto": "✅ 可自動偵測",
    "partial": "⚠️ 部分可偵測",
    "manual": "❌ 需人工評閱",
}

_DETECTABLE_COLORS = {
    "auto": "D8F5E1",  # 綠
    "partial": "FFF3CD",  # 黃
    "manual": "FDDEDE",  # 紅
}


def export_to_excel(items: list[RubricItem], summary: str = "") -> bytes:
    """Generate an .xlsx file from RubricItem list, return as bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "評分表"

    # ── Header ──
    header_font = Font(bold=True, size=11)
    headers = [
        "項目編號",
        "評分項目",
        "說明",
        "配分",
        "可偵測性",
        "自動偵測方式",
        "替代建議",
    ]
    col_widths = [10, 25, 40, 8, 18, 35, 35]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="D0D0D0")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 22

    # ── Data rows ──
    for row_idx, item in enumerate(items, start=2):
        detectable = item.detectable or "manual"
        label = _DETECTABLE_LABELS.get(detectable, detectable)
        bg_color = _DETECTABLE_COLORS.get(detectable, "FFFFFF")
        fill = PatternFill("solid", fgColor=bg_color)

        values = [
            item.id,
            item.title,
            item.description,
            item.max_score,
            label,
            item.detection_method or "",
            item.fallback or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 40

    # ── Summary row ──
    if summary:
        last_row = len(items) + 3
        ws.cell(row=last_row, column=1, value="備註").font = Font(bold=True)
        summary_cell = ws.cell(row=last_row, column=2, value=summary)
        ws.merge_cells(
            start_row=last_row,
            start_column=2,
            end_row=last_row,
            end_column=len(headers),
        )
        summary_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[last_row].height = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
